from odoo import models, fields, api
from odoo.exceptions import UserError
from ....addons import product
from ....addons import stock
from datetime import datetime

class ShopeeOrder(models.Model):
    _name = "shopee.order"
    _description = "An instance of a unique order made on shopee."

    name = fields.Datetime("Invoice Date", default=datetime.now(), readonly=True)
    shopee_id = fields.Many2one("shopee.map", string="Shopee SKU", ondelete="set null")
    quantity = fields.Integer("Quantity", required=True, default=1)
    total_amount = fields.Integer("Total Amount", required=True)

    @api.model
    def create(self, vals):
        order = super(ShopeeOrder, self).create(vals)
        map_entry = self.env['shopee.map'].search([('name', '=', order.shopee_id.name)], limit=1)
        if map_entry:
            product = map_entry.product_id
            stock_quant = self.env['stock.quant'].search([('product_id', '=', product.id)], limit=1)
            if stock_quant:
                stock_quant.ensure_one()
                stock_quant.quantity -= order.quantity

                stock_quant.sudo().write({'quantity': stock_quant.quantity})
        return order
    
    @api.model
    def import_shopee_orders(self, file_path):
        import openpyxl

        # Load the workbook and select the first sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        row_count = 0
        error_count = 1
        error_messages = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_count += 1
            sku = row[6]
            quantity = row[18]
            total_amount = row[19]

            if sku == "-":
                continue

            if not sku or sku == "":
                error_count += 1
                error_messages.append(f">> ERROR {error_count}: SKU in Row {row_count} is empty or invalid.")

            if quantity == "":
                error_count += 1
                error_messages.append(f">> ERROR {error_count}: Quantity in Row {row_count} is empty or invalid.")

            if total_amount == "":
                error_count += 1
                error_messages.append(f">> ERROR {error_count}: Total Amount in Row {row_count} is empty or invalid.")

            map_entry = self.env['shopee.map'].search([('name', '=', sku)], limit=1)
            if not map_entry:
                error_count += 1
                error_messages.append(f"SKU '{sku}' not found in Shopee map. Please create a new entry linking Tokpedia SKU '{sku}' with a product in Odoo's inventory.")

        if len(error_messages) != 0:
            error_output = "\n\n".join(error_messages)
            error_output = f"{error_output}\n\nDatabase operation has been ABORTED due to the above {error_count} errors."
            raise UserError(error_output)

        # Iterate through rows in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Assuming the SKU is in the first column
            sku = row[6]
            # Assuming the quantity is in the third column
            quantity = row[18]
            # Assuming the total amount is in the fourth column
            total_amount = row[19]

            # Skip rows without SKU
            if sku == "-" or quantity == 0 or total_amount == 0:
                continue

            # Find the shopee map entry based on SKU
            map_entry = self.env['shopee.map'].search([('name', '=', sku)], limit=1)

            # Create the order record
            order_vals = {
                'name': datetime.now(),
                'shopee_id': map_entry.id,
                'quantity': quantity,
                'total_amount': total_amount,
            }
            self.create(order_vals)

        return True