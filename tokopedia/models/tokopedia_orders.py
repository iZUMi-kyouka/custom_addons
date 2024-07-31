from odoo import models, fields, api
from odoo.exceptions import UserError
from ....addons import product
from ....addons import stock
from datetime import datetime

class TokopediaOrder(models.Model):
    _name = "tokopedia.order"
    _description = "An instance of a unique order made on Tokopedia."

    name = fields.Char("Invoice No.", required=True)
    # payment_date = fields.Date("Payment Date", required=True)
    tokopedia_id = fields.Many2one("tokopedia.map", string="Tokopedia SKU", ondelete="set null")
    quantity = fields.Integer("Quantity", required=True, default=1)
    total_amount = fields.Integer("Total Amount", required=True)

    @api.model
    def create(self, vals):
        order = super(TokopediaOrder, self).create(vals)
        map_entry = self.env['tokopedia.map'].search([('name', '=', order.tokopedia_id.name)], limit=1)
        if map_entry:
            product = map_entry.product_id
            stock_quant = self.env['stock.quant'].search([('product_id', '=', product.id)], limit=1)
            if stock_quant:
                stock_quant.ensure_one()
                stock_quant.quantity -= order.quantity

                stock_quant.sudo().write({'quantity': stock_quant.quantity})
        return order

    @api.model
    def import_tokopedia_orders(self, file_path):
        import openpyxl

        # Load the workbook and select the first sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        row_count = 0
        error_count = 1
        error_messages = []
        for row in sheet.iter_rows(min_row=6, values_only=True):
            row_count += 1
            # Assuming the SKU is in the first column
            invoice_no = row[1]
            # payment_date = row[2]
            order_status = row[3]
            sku = row[10]
            # Assuming the quantity is in the third column
            quantity = row[13]
            # Assuming the total amount is in the fourth column
            selling_price = row[14]
            tokopedia_sku = row[10]

            if order_status == "Dibatalkan Sistem":
                continue

            if invoice_no == "" or not invoice_no :
                error_messages.append(f">> ERROR {error_count}: Invoice No. at Row {row_count} is empty or invalid")
                error_count += 1
                # raise UserError(f"Invoice No. at Row {row_count} is empty or invalid")
            if order_status == "" or not order_status:
                error_messages.append(f">> ERROR {error_count}: Order status at Row {row_count} is empty or invalid")
                error_count += 1
                # raise UserError(f"Order Status at Row {row_count} is empty or invalid")
            if sku == "" or not sku:
                error_messages.append(f">> ERROR {error_count}: SKU at Row {row_count} is empty or invalid")
                error_count += 1
                # raise UserError(f"SKU at Row {row_count} is empty or invalid")
            if quantity == "" or not quantity:
                error_messages.append(f">> ERROR {error_count}: Quantity at Row {row_count} is empty or invalid")
                error_count += 1
                # raise UserError(f"Quantity at Row {row_count} is empty or invalid")
            if selling_price == "" or not selling_price:
                error_messages.append(f">> ERROR {error_count}: Selling Price at Row {row_count} is empty or invalid")
                error_count += 1
                # raise UserError(f"Selling Price at Row {row_count} is empty or invalid")
            
            quantity = int(quantity)
            total_amount = int(selling_price) * quantity

            tokopedia_sku_map_entry = self.env['tokopedia.map'].search([('name', '=', tokopedia_sku)], limit=1)
            if not tokopedia_sku_map_entry:
                error_messages.append(f">> ERROR {error_count}: SKU '{tokopedia_sku}' not found in Tokopedia map. Please create a new entry linking Tokpedia SKU '{tokopedia_sku}' with a product in Odoo's inventory.")
                error_count += 1
                # raise UserError(f"Failed to import orders. SKU  '{tokopedia_sku}' not found in Tokopedia map. Please create a new entry linking Tokpedia SKU '{tokopedia_sku}' with a product in Odoo's inventory.\n Operation has been ABORTED.")

        if not len(error_messages) == 0:
            error_output = "\n\n".join(error_messages)
            error_output = f"{error_output}\n\nDatabase operation has been ABORTED due to {error_count-1} errors above."
            raise UserError(error_output)

        # Iterate through rows in the sheet
        for row in sheet.iter_rows(min_row=6, values_only=True):
            # Assuming the SKU is in the first column
            invoice_no = row[1]
            # payment_date = row[2]
            order_status = row[3]
            sku = row[10]
            # Assuming the quantity is in the third column
            quantity = int(row[13])
            # Assuming the total amount is in the fourth column
            total_amount = quantity * int(row[14])

            # Skip rows without SKU
            if sku == "" or quantity == 0 or total_amount == 0 or order_status == "Dibatalkan Sistem":
                continue
            
            map_entry = self.env['tokopedia.map'].search([('name', '=', sku)], limit=1)
            
            # Create the order record
            order_vals = {
                'name': invoice_no, 
                'tokopedia_id': map_entry.id,
                'quantity': quantity,
                'total_amount': total_amount,
            }
            self.create(order_vals)

        return True